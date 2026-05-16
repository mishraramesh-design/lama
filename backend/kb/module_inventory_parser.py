"""Generic Module Inventory Parser for LAMA.

Supports: Excel (.xlsx), CSV (.csv), JSON (.json)
Works with any legacy application regardless of language or framework.

Column detection is fuzzy — the parser looks for columns whose names
CONTAIN certain keywords, not exact matches. This makes it resilient to:
  "Module Name" / "Module" / "Business Module" / "Domain"
  "Controller" / "Service" / "Handler" / "Class" / "Component"
  "Tables" / "DB Tables" / "Referenced Tables" / "Database Objects"
"""
import io
import json
import re
from typing import List, Dict, Any, Optional, Tuple


# ── Column detection helpers ──────────────────────────────────────────
def _find_col(header: List[str], *keywords: str) -> Optional[int]:
    """Return index of first column whose name contains any keyword (case-insensitive)."""
    kw_lower = [k.lower() for k in keywords]
    for i, h in enumerate(header):
        h_lower = (h or "").lower().strip()
        if any(k in h_lower for k in kw_lower):
            return i
    return None


def _parse_list(value: Any) -> List[str]:
    """Parse a comma/semicolon/pipe/newline-delimited value into a clean list."""
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if v]
    raw = str(value).strip()
    if not raw:
        return []
    parts = re.split(r"[,;\|\n]+", raw)
    return [p.strip() for p in parts if p.strip()]


def _to_int(value: Any) -> int:
    try:
        return int(str(value).strip())
    except Exception:
        return 0


# ── Excel parser ──────────────────────────────────────────────────────
def _parse_excel(content: bytes) -> List[Dict[str, Any]]:
    """Parse .xlsx — auto-detects module-summary + component-detail sheets, or single combined sheet."""
    try:
        import openpyxl
    except ImportError:
        return []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return []

    sheet_names = wb.sheetnames
    if not sheet_names:
        return []

    module_map: Dict[str, Dict] = {}

    def read_sheet_rows(ws) -> Tuple[List[str], List[Dict]]:
        rows = []
        header = None
        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = [str(c).strip() if c is not None else "" for c in row]
                continue
            if not any(row):
                continue
            rows.append(dict(zip(header, row)))
        return header or [], rows

    # Identify sheets by content
    module_sheet = None
    component_sheet = None
    for name in sheet_names:
        ws = wb[name]
        hdr, _ = read_sheet_rows(ws)
        has_module = _find_col(hdr, "module", "domain", "area", "feature", "package") is not None
        has_component = _find_col(hdr, "controller", "service", "handler", "class", "component", "bean", "facade", "file") is not None
        has_tables = _find_col(hdr, "table", "db", "database", "entity", "model") is not None
        if has_module and not has_component:
            module_sheet = name
        elif has_component and has_tables:
            component_sheet = name
        elif has_module and has_component:
            module_sheet = name
            component_sheet = name
    if module_sheet is None:
        module_sheet = sheet_names[0]
    if component_sheet is None:
        component_sheet = sheet_names[0] if len(sheet_names) == 1 else sheet_names[-1]

    # Parse module summary sheet
    ws = wb[module_sheet]
    hdr, rows = read_sheet_rows(ws)
    col_module = _find_col(hdr, "module", "domain", "area", "feature", "package", "subsystem")
    col_comp_cnt = _find_col(hdr, "controller", "service", "class", "component", "count", "number")
    col_tbl_cnt = _find_col(hdr, "table", "db", "database", "entity", "reference", "estimated")
    col_sample = _find_col(hdr, "sample", "example", "key table", "list")
    col_desc = _find_col(hdr, "description", "detail", "note", "comment", "remark")

    if col_module is not None:
        for row_d in rows:
            vals = list(row_d.values())
            mod_name = str(vals[col_module] or "").strip() if col_module < len(vals) else ""
            if not mod_name:
                continue
            comp_cnt = _to_int(vals[col_comp_cnt]) if col_comp_cnt is not None and col_comp_cnt < len(vals) else 0
            tbl_cnt = _to_int(vals[col_tbl_cnt]) if col_tbl_cnt is not None and col_tbl_cnt < len(vals) else 0
            sample = _parse_list(vals[col_sample]) if col_sample is not None and col_sample < len(vals) else []
            desc = str(vals[col_desc] or "").strip() if col_desc is not None and col_desc < len(vals) else ""
            module_map[mod_name] = {
                "type": "MODULE",
                "name": mod_name,
                "source": "user_import",
                "source_format": "excel",
                "component_count": comp_cnt,
                "table_ref_count": tbl_cnt,
                "sample_tables": sample,
                "components": [],
                "components_detail": [],
                "tables": list(sample),
                "description": desc,
            }

    # Parse component detail sheet
    if component_sheet != module_sheet:
        ws2 = wb[component_sheet]
        hdr2, rows2 = read_sheet_rows(ws2)
    else:
        hdr2, rows2 = hdr, rows

    col2_module = _find_col(hdr2, "module", "domain", "area", "feature", "package", "subsystem")
    col2_name = _find_col(hdr2, "controller", "service", "handler", "class", "component", "bean", "facade", "file", "name")
    col2_path = _find_col(hdr2, "path", "file", "location", "relative", "src", "source")
    col2_tables = _find_col(hdr2, "referenced table", "db table", "table", "entity", "model", "database object")
    col2_tcount = _find_col(hdr2, "table count", "count", "number", "referenced tables count", "table ref")

    comp_entities = []
    if col2_name is not None:
        for row_d in rows2:
            vals2 = list(row_d.values())

            def _v(col):
                return vals2[col] if col is not None and col < len(vals2) else None

            mod_name = str(_v(col2_module) or "").strip()
            comp_name = str(_v(col2_name) or "").strip()
            file_path = str(_v(col2_path) or "").strip()
            tables = _parse_list(_v(col2_tables))
            tbl_count = _to_int(_v(col2_tcount)) if _v(col2_tcount) else len(tables)

            if not comp_name:
                continue
            comp_detail = {"name": comp_name, "file_path": file_path, "tables": tables, "table_count": tbl_count}

            if mod_name and mod_name not in module_map:
                module_map[mod_name] = {
                    "type": "MODULE",
                    "name": mod_name,
                    "source": "user_import",
                    "source_format": "excel",
                    "component_count": 0,
                    "table_ref_count": 0,
                    "sample_tables": [],
                    "components": [],
                    "components_detail": [],
                    "tables": [],
                    "description": "",
                }
            if mod_name in module_map:
                module_map[mod_name]["components"].append(comp_name)
                module_map[mod_name]["components_detail"].append(comp_detail)
                existing = set(module_map[mod_name]["tables"])
                existing.update(tables)
                module_map[mod_name]["tables"] = list(existing)

            comp_entities.append({
                "type": "COMPONENT_MAP",
                "name": comp_name,
                "module": mod_name,
                "file_path": file_path,
                "tables": tables,
                "table_count": tbl_count,
            })

    # Final counts
    for m in module_map.values():
        if m["component_count"] == 0 and m["components"]:
            m["component_count"] = len(m["components"])
        if m["table_ref_count"] == 0 and m["tables"]:
            m["table_ref_count"] = len(m["tables"])

    return list(module_map.values()) + comp_entities


# ── CSV parser ────────────────────────────────────────────────────────
def _parse_csv(content: bytes) -> List[Dict[str, Any]]:
    import csv
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return []
    header = [str(f).strip() for f in reader.fieldnames]
    rows = list(reader)

    col_module = _find_col(header, "module", "domain", "area", "feature", "package", "subsystem")
    col_comp = _find_col(header, "controller", "service", "handler", "class", "component", "bean", "facade", "file", "name")
    col_path = _find_col(header, "path", "file", "location", "relative", "src", "source")
    col_tables = _find_col(header, "referenced table", "db table", "table", "entity", "model")
    col_tcount = _find_col(header, "table count", "count", "referenced tables count", "table ref")
    col_desc = _find_col(header, "description", "detail", "note", "comment")

    module_map: Dict[str, Dict] = {}
    comp_entities: List[Dict] = []

    for row in rows:
        vals = list(row.values())

        def _v(col):
            return vals[col] if col is not None and col < len(vals) else None

        mod_name = str(_v(col_module) or "").strip()
        comp_name = str(_v(col_comp) or "").strip()
        file_path = str(_v(col_path) or "").strip()
        tables = _parse_list(_v(col_tables))
        tbl_count = _to_int(_v(col_tcount)) if _v(col_tcount) else len(tables)
        desc = str(_v(col_desc) or "").strip()

        if mod_name and mod_name not in module_map:
            module_map[mod_name] = {
                "type": "MODULE",
                "name": mod_name,
                "source": "user_import",
                "source_format": "csv",
                "component_count": 0,
                "table_ref_count": 0,
                "sample_tables": tables[:5],
                "components": [],
                "components_detail": [],
                "tables": [],
                "description": desc,
            }
        comp_detail = {
            "name": comp_name or mod_name,
            "file_path": file_path,
            "tables": tables,
            "table_count": tbl_count,
        }
        if mod_name in module_map:
            if comp_name:
                module_map[mod_name]["components"].append(comp_name)
                module_map[mod_name]["components_detail"].append(comp_detail)
            existing = set(module_map[mod_name]["tables"])
            existing.update(tables)
            module_map[mod_name]["tables"] = list(existing)

        if comp_name:
            comp_entities.append({
                "type": "COMPONENT_MAP",
                "name": comp_name,
                "module": mod_name,
                "file_path": file_path,
                "tables": tables,
                "table_count": tbl_count,
            })

    for m in module_map.values():
        m["component_count"] = len(m["components"])
        m["table_ref_count"] = len(m["tables"])

    return list(module_map.values()) + comp_entities


# ── JSON parser ───────────────────────────────────────────────────────
def _parse_json(content: bytes) -> List[Dict[str, Any]]:
    try:
        data = json.loads(content.decode("utf-8-sig", errors="replace"))
    except Exception:
        return []

    entities: List[Dict] = []
    module_map: Dict[str, Dict] = {}

    raw_modules: List[Dict] = []
    if isinstance(data, list):
        raw_modules = data
    elif isinstance(data, dict):
        for key in ("modules", "data", "items", "inventory", "result"):
            if key in data and isinstance(data[key], list):
                raw_modules = data[key]
                break
        if not raw_modules:
            for k, v in data.items():
                if isinstance(v, dict):
                    raw_modules.append({"name": k, **v})
                elif isinstance(v, list):
                    raw_modules.append({"name": k, "components": v})

    for item in raw_modules:
        if not isinstance(item, dict):
            continue
        name = (item.get("name") or item.get("module") or item.get("domain") or item.get("area") or item.get("package") or "")
        name = str(name).strip()
        if not name:
            continue
        desc = str(item.get("description") or item.get("desc") or item.get("note") or "").strip()
        raw_comps = (item.get("components") or item.get("controllers") or item.get("services") or item.get("classes") or item.get("handlers") or item.get("files") or [])

        components_detail = []
        all_tables: set = set()
        if isinstance(raw_comps, list):
            for c in raw_comps:
                if isinstance(c, str):
                    components_detail.append({"name": c, "file_path": "", "tables": [], "table_count": 0})
                elif isinstance(c, dict):
                    cname = str(c.get("name") or c.get("class") or c.get("file") or c.get("controller") or c.get("service") or "").strip()
                    cpath = str(c.get("path") or c.get("file_path") or c.get("file") or "").strip()
                    ctbls = _parse_list(c.get("tables") or c.get("db_tables") or c.get("entities") or c.get("references") or [])
                    components_detail.append({
                        "name": cname,
                        "file_path": cpath,
                        "tables": ctbls,
                        "table_count": len(ctbls),
                    })
                    all_tables.update(ctbls)

        mod_tables = _parse_list(item.get("tables") or item.get("db_tables") or item.get("entities") or [])
        all_tables.update(mod_tables)

        module_map[name] = {
            "type": "MODULE",
            "name": name,
            "source": "user_import",
            "source_format": "json",
            "component_count": len(components_detail),
            "table_ref_count": len(all_tables),
            "sample_tables": list(all_tables)[:10],
            "components": [c["name"] for c in components_detail],
            "components_detail": components_detail,
            "tables": list(all_tables),
            "description": desc,
        }
        for cd in components_detail:
            entities.append({
                "type": "COMPONENT_MAP",
                "name": cd["name"],
                "module": name,
                "file_path": cd["file_path"],
                "tables": cd["tables"],
                "table_count": cd["table_count"],
            })

    return list(module_map.values()) + entities


# ── Public API ────────────────────────────────────────────────────────
def parse_module_inventory(filename: str, content: bytes) -> List[Dict[str, Any]]:
    """Auto-detect format from extension and parse.

    Returns list of MODULE + COMPONENT_MAP entities. Returns [] on any failure.
    """
    name = (filename or "").lower()
    try:
        if name.endswith(".xlsx") or name.endswith(".xls"):
            return _parse_excel(content)
        elif name.endswith(".csv"):
            return _parse_csv(content)
        elif name.endswith(".json"):
            return _parse_json(content)
        else:
            return _parse_excel(content) or _parse_csv(content) or _parse_json(content)
    except Exception:
        return []


def generate_module_text_summary(entities: List[Dict[str, Any]]) -> str:
    """Human-readable summary of module inventory for RAG indexing."""
    modules = [e for e in entities if e.get("type") == "MODULE"]
    if not modules:
        return ""

    lines = [
        "MODULE INVENTORY — LEGACY APPLICATION STRUCTURE",
        f"Total modules identified: {len(modules)}",
        f"Total components mapped: {sum(m.get('component_count', 0) for m in modules)}",
        f"Total table references: {sum(m.get('table_ref_count', 0) for m in modules)}",
        "",
    ]
    for m in sorted(modules, key=lambda x: x.get("table_ref_count", 0), reverse=True):
        lines.append(f"Module: {m['name']} | Components: {m.get('component_count', 0)} | Table refs: {m.get('table_ref_count', 0)}")
        if m.get("description"):
            lines.append(f"  Description: {m['description'][:200]}")
        if m.get("tables"):
            lines.append(f"  Key tables: {', '.join(m['tables'][:12])}")
        if m.get("components"):
            extra = f" +{len(m['components']) - 8} more" if len(m['components']) > 8 else ""
            lines.append(f"  Components: {', '.join(m['components'][:8])}{extra}")
        lines.append("")
    return "\n".join(lines)
